import os
import django
import openpyxl

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "analysis_tool.settings")
django.setup()

# 🚀 Import modeli przenosimy tu:


def run_import():
    from surveys.models import Answer  # ← to działa, edytor nie ruszy!

    workbook = openpyxl.load_workbook("dane.xlsx")

    def extract_value_and_type(sheet_name):
        parts = sheet_name.strip().split()
        if len(parts) < 2:
            return None, None
        typ_raw = parts[-1].lower()
        typ = "skojarzenie" if "skoj" in typ_raw else "definicja"
        # łączy wszystkie części oprócz ostatniej
        value = " ".join(parts[:-1]).strip().upper()
        return value, typ

    for sheet_name in workbook.sheetnames:
        print(f"📄 Przetwarzam arkusz: {sheet_name}")
        value, typ = extract_value_and_type(sheet_name)
        if not value:
            print(f"⚠️ Pomijam arkusz: {sheet_name}")
            continue

        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            content_pl = row[5]
            if isinstance(content_pl, str) and content_pl.strip():
                Answer.objects.create(
                    value=value,
                    language="pl",
                    type=typ,
                    content=content_pl.strip()
                )

            content_ua = row[0]
            if isinstance(content_ua, str) and content_ua.strip():
                Answer.objects.create(
                    value=value,
                    language="ua",
                    type=typ,
                    content=content_ua.strip()
                )

    print("✅ Import zakończony!")


# 👇 wywołanie
run_import()
